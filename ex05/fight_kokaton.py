import webbrowser
import pygame as pg
import random
import openpyxl
import pandas as pd
import time
import sys


class Screen:
    def __init__(self, title, wh, img_path):
        # 練習1
        pg.display.set_caption(title)
        self.sfc = pg.display.set_mode(wh)
        self.rct = self.sfc.get_rect()
        self.bgi_sfc = pg.image.load(img_path)
        self.bgi_rct = self.bgi_sfc.get_rect()

    def blit(self):
        self.sfc.blit(self.bgi_sfc, self.bgi_rct)


class Bird:
    key_delta = {
        pg.K_UP:    [0, -1],
        pg.K_DOWN:  [0, +1],
        pg.K_LEFT:  [-1, 0],
        pg.K_RIGHT: [+1, 0],
    }

    def __init__(self, img_path, ratio, xy):
        self.sfc = pg.image.load(img_path)
        self.sfc = pg.transform.rotozoom(self.sfc, 0, ratio)
        self.rct = self.sfc.get_rect()
        self.rct.center = xy

    def blit(self, scr: Screen):
        scr.sfc.blit(self.sfc, self.rct)

    def update(self, scr):
        key_dct = pg.key.get_pressed()
        for key, delta in Bird.key_delta.items():
            if key_dct[key]:
                self.rct.centerx += delta[0]
                self.rct.centery += delta[1]
            if check_bound(self.rct, scr.rct) != (+1, +1):
                self.rct.centerx -= delta[0]
                self.rct.centery -= delta[1]
        self.blit(scr)


class Bomb:
    def __init__(self, color, rad, vxy, scr: Screen):
        self.sfc = pg.Surface((2*rad, 2*rad))  # 正方形の空のSurface
        self.sfc.set_colorkey((0, 0, 0))
        pg.draw.circle(self.sfc, color, (rad, rad), rad)
        self.rct = self.sfc.get_rect()
        self.rct.centerx = random.randint(0, scr.rct.width)
        self.rct.centery = random.randint(0, scr.rct.height)
        self.vx, self.vy = vxy

    def blit(self, scr: Screen):
        scr.sfc.blit(self.sfc, self.rct)

    def update(self, scr: Screen):
        self.rct.move_ip(self.vx, self.vy)
        self.sfc.blit(self.sfc, self.rct)
        yoko, tate = check_bound(self.rct, scr.rct)
        self.vx *= yoko
        self.vy *= tate
        self.blit(scr)


# スコアブックを生成する。（Excelファイル）

class Commonfile:
    def __init__(self, filename):
        self.filename = filename


class ExcelScore(Commonfile):

    def generate_file(self, value):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "経過時間"
        ws["B1"] = value
        wb.save(self.filename)


class Csv(Commonfile):
    def generate_file(self):
        df = pd.read_excel(self.filename)
        df.to_csv("ex05/score.csv", index=False)


def check_bound(obj_rct, scr_rct):
    """
    第1引数：こうかとんrectまたは爆弾rect
    第2引数：スクリーンrect
    範囲内：+1／範囲外：-1
    """
    yoko, tate = +1, +1
    if obj_rct.left < scr_rct.left or scr_rct.right < obj_rct.right:
        yoko = -1
    if obj_rct.top < scr_rct.top or scr_rct.bottom < obj_rct.bottom:
        tate = -1
    return yoko, tate


def main():
    global time_score
    clock = pg.time.Clock()
    start = time.time()

    # 練習1
    scr = Screen("負けるなこうかとん!", (1600, 900), "fig/pg_bg.jpg")

    # 練習３
    # こうかとんの画像集
    kkt_images = [f"fig/{i}.png" for i in range(0, 10)]
    # こうかとんの画像を無作為に選択する。
    kkt = Bird(kkt_images[random.randint(0, 10)], 2.0, (900, 400))
    kkt.update(scr)

    # 練習５
    bkd_lst = []
    # ボールがランダムに数が変わるように変数を定義
    rdm_boll = random.randint(1, 6)
    for _ in range(rdm_boll):
        # 追加機能:ボールの色が起動するごとに変わる
        # 3原色を管理するタプル（内包表記）
        random_color_tpl = (random.randint(0, 256) for _ in range(3))
        color1, color2, color3 = random_color_tpl
        bkd = Bomb((color1, color2, color3), 10, (1, 1), scr)
        bkd_lst.append(bkd)

    # 練習２
    while True:
        scr.blit()

        for event in pg.event.get():
            if event.type == pg.QUIT:
                return

        kkt.update(scr)
        for i in range(rdm_boll):
            bkd_lst[i].update(scr)
            bkd.update(scr)
            if kkt.rct.colliderect(bkd_lst[i].rct):
                # 学内ポータルサイトが開く
                url = "https://service.cloud.teu.ac.jp/portal/index?"
                webbrowser.open(url)
                return

        pg.display.update()
        clock.tick(1000)
        time_score = time.time()-start


if __name__ == "__main__":
    pg.init()
    main()
    # Excelスコアファイルを出力
    scorefile_excel = ExcelScore("ex05/score1.xlsx")
    scorefile_excel.generate_file(time_score)
    # CSVスコアファイルを出力
    scorefile_csv = Csv("ex05/score1.xlsx")
    scorefile_csv.generate_file()
    pg.quit()
    sys.exit()
